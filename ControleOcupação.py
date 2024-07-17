import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os

# Defina o caminho do arquivo TXT e Excel
caminho_arquivo_txt = r'C:\Users\an770843\Mars Inc\Brazil_MW_PlanningComex - Documents\2 - DRP\2024\PO Intercompany\INPUT_SAP\Stock.txt'
caminho_excel = r'C:\Users\an770843\Mars Inc\Brazil_MW_PlanningComex - Documents\2 - DRP\2024\Controle Ocupação\BaseOcupação.xlsx'

# Defina os nomes das colunas conforme os cabeçalhos do arquivo original
colunas = ['Material', 'Cen.', 'Texto breve de material', 'Dep.', 'Lote', 'UMB', 'Util.livre', 'Em CtrQld.', 'Bloqueado', 'Trâns.e TE', 'Restrito']

# Função para ler e processar o arquivo TXT
def processar_txt(caminho_arquivo):
    with open(caminho_arquivo, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    # Ignorar as primeiras 3 linhas que não contêm dados relevantes
    lines = lines[3:]

    # Inicializar uma lista para armazenar as linhas de dados válidas
    dados_validos = []

    # Filtrar as linhas que seguem o padrão específico
    for line in lines:
        # Remover espaços desnecessários e quebrar a linha nos delimitadores "|"
        partes = [parte.strip() for parte in line.strip().split('|')[1:-1]]
        # Verificar se o número de partes é igual ao número de colunas esperadas
        if len(partes) == len(colunas) and partes != colunas:
            dados_validos.append(partes)

    # Converter a lista de dados válidos em um DataFrame
    df = pd.DataFrame(dados_validos, columns=colunas)

    # Remover pontos dos números maiores que 1000 e converter para int
    for col in ['Util.livre', 'Em CtrQld.', 'Bloqueado', 'Trâns.e TE', 'Restrito']:
        df[col] = df[col].str.replace('.', '')  # Remover pontos
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)  # Converter para numérico

    return df

# Processar o arquivo TXT
df_txt = processar_txt(caminho_arquivo_txt)

# Adicionar a data atual no formato datetime
df_txt['Data'] = pd.to_datetime(datetime.now().strftime('%d/%m/%Y'), format='%d/%m/%Y')

# Verificar se o arquivo Excel já existe
if os.path.exists(caminho_excel):
    # Carregar o workbook existente
    wb = openpyxl.load_workbook(caminho_excel)
    ws = wb.active

    # Converter DataFrame em linhas do Excel
    rows = dataframe_to_rows(df_txt, index=False, header=False)

    # Identificar a última linha preenchida e adicionar os novos dados
    last_row = ws.max_row
    for r_idx, row in enumerate(rows, last_row + 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
else:
    # Se o arquivo não existir, criar um novo workbook e worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BaseOcupação"

    # Escrever o cabeçalho e os dados
    ws.append(colunas + ['Data'])
    for row in dataframe_to_rows(df_txt, index=False, header=False):
        ws.append(row)

# Formatar a coluna "Data" como data sem horas
for cell in ws['L'][1:]:  # Coluna "L" é a 12ª coluna, onde está a coluna "Data"
    cell.number_format = 'DD/MM/YYYY'

# Salvar os dados atualizados no arquivo Excel
wb.save(caminho_excel)

print("Controle de estoque atualizado com sucesso.")